import argparse
import openpyxl
from openpyxl import load_workbook
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import precision_score, roc_auc_score, recall_score, f1_score, confusion_matrix, accuracy_score, \
    matthews_corrcoef, classification_report, auc, roc_curve
import os
from datetime import datetime
import joblib
import numpy as np
import pandas as pd
import matplotlib
from matplotlib import pyplot as plt
from sklearn.ensemble import AdaBoostClassifier, RandomForestClassifier
from sklearn import svm
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import CategoricalNB, GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import label_binarize, StandardScaler
from sklearn.tree import DecisionTreeClassifier
from script.utils import polt_confusion_matrix, plot_curve
matplotlib.use('agg')  # 设置后端为agg

# 计算特异度和阴性预测值
def calculate_specificity_npv(predictions, true_labels, class_label):
    # 将预测结果和真实标签转换为NumPy数组
    y_pred = np.array(predictions)
    y_true = np.array(true_labels)

    # 根据类别标签获取正例和负例的索引
    positive_indices = np.where(y_true == class_label)[0]
    negative_indices = np.where(y_true != class_label)[0]

    # 计算特异度
    true_negatives = np.sum(y_pred[negative_indices] == y_true[negative_indices])
    false_positives = np.sum(y_pred[positive_indices] != y_true[positive_indices])
    specificity = true_negatives / (true_negatives + false_positives)

    # 计算阴性预测值
    false_negatives = np.sum(y_pred[negative_indices] != y_true[negative_indices])
    negative_predictive_value = true_negatives / (true_negatives + false_negatives)

    return specificity, negative_predictive_value


# 定义评价指标的函数
def evaluate_model(model, X, y, number, type='train'):
    # 计算每个类别的AUC
    fpr = dict()
    tpr = dict()
    n_classes = 4
    roc_auc = dict()
    y_pred = model.predict(X)

    # 计算每个类别的Specificity, npv
    specificities, npvs = [], []
    for i in range(len(np.unique(y))):
        sp, npv = calculate_specificity_npv(y_pred, y, i)
        specificities.append(sp)
        npvs.append(npv)
    # for i in range(len(np.unique(y))):
    #     cm = confusion_matrix(y, y_pred, labels=[i])
    #     tn = cm.sum() - cm.sum(axis=0)[0] - cm.sum(axis=1)[0] + cm[0][0]
    #     fp = cm.sum(axis=0)[0] - cm[0][0]
    #     fn = cm.sum(axis=1)[0] - cm[0][0]
    #     specificity = tn / (tn + fp)
    #     npv = tn / (tn + fn)
    #     specificities.append(specificity)
    #     npvs.append(npv)
    # 计算recall
    report = classification_report(y, y_pred, output_dict=True)
    all_sensitivity = dict()
    for i in range(4):
        all_sensitivity[str(i)] = report[str(i)]['recall']

    probability = model.predict_proba(X)
    # 在这里更新全局数组，记录全部预测和真实标签，以便计算roc
    if type == 'test':
        global all_predict, all_true_y
        if all_predict is None:
            all_predict = probability
        else:
            all_predict = np.concatenate((all_predict, probability), axis=0)
        all_true_y = np.append(all_true_y, y)
    accuracy = accuracy_score(y, y_pred)  # 准确率
    sensitivity = recall_score(y, y_pred, average='macro')  # 敏感性
    # recall
    for i in range(n_classes):
        fpr[i], tpr[i], _ = roc_curve(y == i, probability[:, i])
        roc_auc[i] = roc_auc_score(y == i, probability[:, i])
    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))
    mean_tpr = np.zeros_like(all_fpr)
    for i in range(n_classes):
        mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])
    mean_tpr /= n_classes
    fpr["macro"] = all_fpr
    tpr["macro"] = mean_tpr
    roc_auc['macro'] = auc(fpr['macro'], tpr['macro'])

    precision = precision_score(y, y_pred, labels=np.arange(number), average='weighted')  # 精准度
    f1 = f1_score(y, y_pred, average='weighted')  # F1
    mcc = matthews_corrcoef(y, y_pred)  # 皮尔森相关系数
    return accuracy, sensitivity, np.mean(specificities), roc_auc, precision, f1, mcc, np.mean(npvs), y_pred, fpr, tpr,

def train(file_name, model_type, classfication_num):
    best_acc = -1
    df = pd.read_csv(file_name, header=0)
    data = df.replace({'DHP': 0, 'PG': 1, 'XH': 2, 'ZTH': 3}).values
    # 选择特定列数据
    selected_columns = [4, 9, 10, 11]
    features = df.iloc[:, selected_columns].values

    # features = data[::, 2::]
    labels = data[::, 1].astype(np.int32)
    # 标准化
    scaler_standard = StandardScaler()
    features = scaler_standard.fit_transform(features)
    # labels = label_binarize(labels, classes=[0, 1, 2, 3])
    # 获取当前日期,为后面模型添加时间戳
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    # 结果保存
    save_path = '../result4/' + model_type + date + 'result.xlsx'
    if not os.path.exists(save_path):
        df = pd.DataFrame()  # 表示创建空的表
        df.to_excel(save_path)

    train_acc = []
    epoch = 100
    confusion_matrix_ = np.zeros((classfication_num, classfication_num))
    confusion_matrix_.tolist()
    t_confusion_matrix_ = np.zeros((classfication_num, classfication_num))
    t_confusion_matrix_.tolist()
    # 各项指标
    all_ac, all_ps, all_rs, all_fs, all_sp, all_roc, all_mcc, all_npv = [], [], [], [], [], [], [], []
    t_all_ac, t_all_ps, t_all_rs, t_all_fs, t_all_sp, t_all_roc, t_all_mcc, t_all_npv = [], [], [], [], [], [], [], []
    # 根据条件选择要使用的机器学习算法
    # 逻辑回归 (Logistic Regression)
    if model_type == "LogisticRegression":
        model = LogisticRegression(solver='newton-cg', C=2.0)
    # 支持向量机 (Support Vector Machine)
    elif model_type == "SVM":
        model = svm.SVC(C=1.0, probability=True, kernel='linear')
    # K最近邻 (K-Nearest Neighbors)
    elif model_type == "KNN":
        model = KNeighborsClassifier(n_neighbors=5)
    # XGBoost
    elif model_type == "AdaBoost":
        model = AdaBoostClassifier(DecisionTreeClassifier(max_depth=3, min_samples_split=20, min_samples_leaf=5),
                         algorithm="SAMME",
                         n_estimators=30, learning_rate=0.8)
    # 随机森林 (Random Forest)
    elif model_type == "RandomForest":
        model = RandomForestClassifier(n_estimators=50)
    # 线性判别分析 (Linear Discriminant Analysis)
    elif model_type == "LDA":
        model = LinearDiscriminantAnalysis()
    # 二次判别分析 (Quadratic Discriminant Analysis)
    elif model_type == "QDA":
        model = QuadraticDiscriminantAnalysis()
    # 人工神经网络 (Artificial Neural Network)
    elif model_type == "ANN":
        model = MLPClassifier()
    # 朴素贝叶斯
    elif model_type == "naive_bayes":
        model = GaussianNB()
    else:
        raise ValueError("Invalid condition specified")

    for i in range(epoch):
        # 数据集加载
        train_features, test_features, train_labels, test_labels = train_test_split(features, labels, test_size=0.23,
                                                                                    stratify=labels, random_state=i)
        # training the model
        model.fit(train_features, train_labels)
        train_score = model.score(train_features, train_labels)
        train_acc.append(train_score)

        print('Start predicting...')
        accuracy, sensitivity, specificity, auc, precision, f1, mcc, npv, test_predict, fpr, tpr = evaluate_model(model, test_features,
                                                                                         test_labels, number=classfication_num, type='test')
        t_accuracy, t_sensitivity, t_specificity, t_auc, t_precision, t_f1, t_mcc, t_npv, t_test_predict, t_fpr, t_tpr = evaluate_model(model,
                                                                                                                  train_features,
                                                                                                                  train_labels,
                                                                                                                  number=classfication_num)
        if accuracy > best_acc:
            best_acc = accuracy
            joblib.dump(model, "../models/" + model_type + '.m')
        # 记录指标
        all_ps.append(precision)
        t_all_ps.append(t_precision)
        all_rs.append(sensitivity)
        t_all_rs.append(t_sensitivity)
        all_fs.append(f1)
        t_all_fs.append(t_f1)
        all_sp.append(specificity)
        t_all_sp.append(t_specificity)
        all_roc.append(auc['macro'])
        t_all_roc.append(t_auc['macro'])
        all_mcc.append(mcc)
        t_all_mcc.append(t_mcc)
        all_ac.append(accuracy)
        t_all_ac.append(t_accuracy)
        all_npv.append(npv)
        t_all_npv.append(t_npv)
        # 分类报告
        # class_report = classification_report(test_labels, test_predict,
        #                                              target_names=['DHP', 'PG', 'XH', 'ZTH'])
        # print(class_report)
        # 输出混淆矩阵
        confusion_matrix_ += confusion_matrix(test_labels, test_predict)
        t_confusion_matrix_ += confusion_matrix(train_labels, t_test_predict)
        print('--混淆矩阵--')
        print(confusion_matrix_)
        print('--训练混淆矩阵--')
        print(t_confusion_matrix_)

    ps_mean = np.mean(all_ps)
    ps_std = np.std(all_ps, ddof=1)
    t_ps_mean = np.mean(t_all_ps)
    t_ps_std = np.std(t_all_ps, ddof=1)
    rs_mean = np.mean(all_rs)
    rs_std = np.std(all_rs, ddof=1)
    t_rs_mean = np.mean(t_all_rs)
    t_rs_std = np.std(t_all_rs, ddof=1)
    fs_mean = np.mean(all_fs)
    fs_std = np.std(all_fs, ddof=1)
    t_fs_mean = np.mean(t_all_fs)
    t_fs_std = np.std(t_all_fs, ddof=1)
    sp_mean = np.mean(all_sp)
    sp_std = np.std(all_sp, ddof=1)
    t_sp_mean = np.mean(t_all_sp)
    t_sp_std = np.std(t_all_sp, ddof=1)
    ac_mean = np.mean(all_ac)
    ac_std = np.std(all_ac, ddof=1)
    t_ac_mean = np.mean(t_all_ac)
    t_ac_std = np.std(t_all_ac, ddof=1)
    roc_mean = np.mean(all_roc)
    roc_std = np.std(all_roc, ddof=1)
    t_roc_mean = np.mean(t_all_roc)
    t_roc_std = np.std(t_all_roc, ddof=1)
    mcc_mean = np.mean(all_mcc)
    mcc_std = np.std(all_mcc, ddof=1)
    t_mcc_mean = np.mean(t_all_mcc)
    t_mcc_std = np.std(t_all_mcc, ddof=1)
    npv_mean = np.mean(all_npv)
    npv_std = np.std(all_npv, ddof=1)
    t_npv_mean = np.mean(t_all_npv)
    t_npv_std = np.std(t_all_npv, ddof=1)
    polt_confusion_matrix(confusion_matrix_/100, ['DHP', 'PG', 'XH', 'ZTH'], model_type,
                          "../result4/confusion_matrix/test/mean" + model_type + date + '.png')
    polt_confusion_matrix(t_confusion_matrix_/100, ['DHP', 'PG', 'XH', 'ZTH'], model_type,
                          "../result4/confusion_matrix/train/mean" + model_type + date + '.png')
    polt_confusion_matrix(confusion_matrix_, ['DHP', 'PG', 'XH', 'ZTH'], model_type,
                          "../result4/confusion_matrix/test/" + model_type + date + '.png')
    polt_confusion_matrix(t_confusion_matrix_, ['DHP', 'PG', 'XH', 'ZTH'], model_type,
                          "../result4/confusion_matrix/train/" + model_type + date + '.png')
    data = {'Acc': str(round(ac_mean, 4)) + '±' + str(round(ac_std, 4)),
            'precision': str(round(ps_mean, 4)) + '±' + str(round(ps_std, 4)),
            'recall': str(round(rs_mean, 4)) + '±' + str(round(rs_std, 4)),
            'F1': str(round(fs_mean, 4)) + '±' + str(round(fs_std, 4)),
            'Specificity': str(round(sp_mean, 4)) + '±' + str(round(sp_std, 4)),
            'AUC': str(round(roc_mean, 4)) + '±' + str(round(roc_std, 4)),
            'MCC': str(round(mcc_mean, 4)) + '±' + str(round(mcc_std, 4)),
            'npv': str(round(npv_mean, 4)) + '±' + str(round(npv_std, 4)),
            }
    data_ = {
        'Acc': str(round(t_ac_mean, 4)) + '±' + str(round(t_ac_std, 4)),
        'precision': str(round(t_ps_mean, 4)) + '±' + str(round(t_ps_std, 4)),
        'recall': str(round(t_rs_mean, 4)) + '±' + str(round(t_rs_std, 4)),
        'F1': str(round(t_fs_mean,4)) + '±' + str(round(t_fs_std, 4)),
        'Specificity': str(round(t_sp_mean, 4)) + '±' + str(round(t_sp_std, 4)),
        'AUC': str(round(t_roc_mean, 4)) + '±' + str(round(t_roc_std, 4)),
        'MCC': str(round(t_mcc_mean, 4)) + '±' + str(round(t_mcc_std, 4)),
        'npv': str(round(t_npv_mean, 4)) + '±' + str(round(t_npv_std, 4)),
    }
    with pd.ExcelWriter(save_path, mode='a', engine='openpyxl') as writer:
        # 删除已存在的 sheet
        if model_type in writer.book.sheetnames:
            del writer.book[model_type]
        if model_type+'_confusion_matrix' in writer.book.sheetnames:
            del writer.book[model_type+'_confusion_matrix']
        df_train = pd.DataFrame(data_, index=[0])
        df_train.to_excel(writer, sheet_name='train' + model_type, index=False)
        df_train_confusion_matrix = pd.DataFrame(t_confusion_matrix_)
        df_train_confusion_matrix.to_excel(writer, sheet_name='train' + model_type + '_confusion_matrix', header=False, index=False)

        df = pd.DataFrame(data, index=[0])
        df.to_excel(writer, sheet_name=model_type, index=False)
        df1 = pd.DataFrame(confusion_matrix_)
        df1.to_excel(writer, sheet_name=model_type + '_confusion_matrix', header=False, index=False)
    sheet_name = 'Sheet1'
    try:
        workbook = openpyxl.load_workbook(save_path)
        worksheet = workbook[sheet_name]
        workbook.remove(worksheet)
        workbook.save(save_path)
    except:
        x = 1
        print('error')

if __name__ == '__main__':
    global all_predict
    global all_true_y
    model_types = ['LogisticRegression', 'SVM', 'KNN', 'AdaBoost', 'RandomForest', 'LDA', 'QDA', 'ANN', 'naive_bayes']
    # model_types = ['RandomForest']
    file_name = '../data/data.csv'
    parser = argparse.ArgumentParser()
    parser.add_argument('--model_types', default=model_types, type=list)
    parser.add_argument('--file_name', default=file_name, type=str)
    parser.add_argument('--classfication_num', default=4, type=int)
    args = parser.parse_args()
    for model_type in args.model_types:
        # 初始化字典来存储每个类别的FPR, TPR和AUC
        all_fpr = {}
        all_tpr = {}
        all_auc = {}
        all_predict = None
        all_true_y = np.array([])
        train(file_name=args.file_name, model_type=model_type, classfication_num=args.classfication_num)
        # 对于每个类别
        for i in range(4):

            # 将当前类标记为正类，其余为负类
            y_true_binary = (all_true_y == i).astype(int)
            # 计算ROC曲线
            fpr, tpr, _ = roc_curve(y_true_binary, all_predict[:, i])
            # 计算AUC
            roc_auc = auc(fpr, tpr)
            # 存储结果
            all_fpr[str(i)] = fpr
            all_tpr[str(i)] = tpr
            all_auc[str(i)] = roc_auc
        macro_fpr = np.unique(np.concatenate([all_fpr[str(i)] for i in range(4)]))
        macro_tpr = np.zeros_like(macro_fpr)
        for i in range(4):
            macro_tpr += np.interp(macro_fpr, all_fpr[str(i)], all_tpr[str(i)])
        macro_tpr /= 4
        all_fpr["macro"] = macro_fpr
        all_tpr["macro"] = macro_tpr
        all_auc['macro'] = auc(all_fpr['macro'], all_tpr['macro'])
        plot_curve(all_fpr, all_tpr, all_auc, model_type, prex='all')
